#!/usr/bin/env python3
"""Build a fail-closed G3-E annotation ingest queue from local artifacts.

This is an authoring utility, not a validator. It performs only targeted
mechanical assertions over the exact Thin_PIT worksheet and emits no admitted
feature values, PASS receipts, or gate claims.
"""

from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path("/workspace/scratch/f56b716343a6")
OLD = Path("/workspace/scratch/577256efb437")
OUT = ROOT / "agent_g3_annotation_candidate"

WORKBOOK = OLD / "qa/wp2_sources/U127_Data_Expansion_Working_v0.8_2026-08-15.xlsx"
BUILD_MANIFEST = OLD / "remediation/r_wp23_data_closure/05_THIN_PIT_SOURCE_ANNOTATION_BUILD_MANIFEST.csv"
WINDOWS = OLD / "remediation/g2_release_candidate_20260824/W1_W8_WINDOW_REGISTRY_RELEASE_CANDIDATE_v0.1.csv"
DENOMINATOR_QUEUE = OLD / "remediation/r_wp23_data_closure/03_DENOMINATOR_CLOSURE_QUEUE.csv"
MEMBERSHIP = OLD / "remediation/g2_release_candidate_20260824/U127_WORKING_MEMBERSHIP_RELEASE_CANDIDATE_v0.1.csv"

OBSERVED_AT_KST = "2026-08-26T01:25:20+09:00"
FIELD_REGISTRY_ID = "AAA-M3TOP3-G3E-ANNOTATION-FIELD-REGISTRY-v0.1"
SCHEMA_ID = "AAA-M3TOP3-G3E-ANNOTATION-INGEST-SCHEMA-v0.1"
QUEUE_ID = "AAA-M3TOP3-G3E-ANNOTATION-INGEST-QUEUE-v0.1"

FEATURE_BLOCKS = [
    "VALUATION",
    "EARNINGS",
    "FORWARD_EXPECTATIONS",
    "GUIDANCE",
    "PO_ORDER",
    "BACKLOG",
    "QUALIFICATION",
    "REPEAT_ORDER",
    "DESIGN_WIN",
    "FAB_CAPEX",
    "MATERIAL_REFS",
]

EXPECTED_FEATURE_FIELDS = [
    "valuation_observation_status",
    "latest_revenue",
    "latest_OP",
    "latest_margin",
    "Forward_EPS",
    "Forward_OP",
    "EPS_OP_revision",
    "consensus_provider",
    "observation_at",
    "guidance",
    "PO_order_status",
    "backlog_status",
    "qualification_status",
    "volume_repeat_order_status",
    "design_win_customer_adoption_status",
    "fab_capex_state_status",
    "latest_material_earnings_guidance_refs",
]

POST_CUTOFF_CONTEXT_DOCS = [
    "01-Samsung_Semiconductor_Future_Technology_Platform_Map_v1.1_2026-08-09.docx",
    "02-Future_Semiconductor_System_Architecture_Memory_Manufacturing_Map_v1.0_2026-08-10.docx",
    "03-Korea_Semiconductor_Equipment_CAPEX_Early_Revenue_Response_Map_v1.0_2026-08-11.docx",
    "04-SK_Hynix_CAPEX_Korea_Equipment_Ecosystem_Report_v0_2026-08-14.docx",
    "05-Korea_Semiconductor_SupplyChain_TOP38_Master_Scorecard_v2.2_Detail_2026-08-14.docx",
    "06-Semi_Eval_Core_v1.0_2026-08-14.docx",
    "07-Semi_Data_Route_v1.1_2026-08-14.docx",
    "08-Semi_Universe_v1.0_2026-08-14.docx",
    "09-SEMI-PIT-LEDGER_v1.0.docx",
    "11-SEMI-SOURCE-INDEX_v1.0.docx",
    "12-SEMI-COMPANY-MASTER_v0.1.docx",
    "14-SEMI-FAB-MASTER_v1.0_2026-08-14.docx",
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def worksheet_rows(wb, title: str) -> tuple[list[str], list[dict[str, Any]]]:
    ws = wb[title]
    headers = [scalar(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row_no in range(2, ws.max_row + 1):
        record = {
            str(headers[col - 1]): scalar(ws.cell(row_no, col).value)
            for col in range(1, len(headers) + 1)
        }
        record["_workbook_row"] = row_no
        rows.append(record)
    return [str(v) for v in headers], rows


def manifest_registry() -> list[dict[str, Any]]:
    rows = read_csv(BUILD_MANIFEST)
    selected = [row for row in rows if row["block"] in FEATURE_BLOCKS]
    registry: list[dict[str, Any]] = []
    for row in selected:
        for field in row["thin_pit_fields"].split("/"):
            registry.append(
                {
                    "field_name": field,
                    "block": row["block"],
                    "current_coverage": row["current_coverage"],
                    "current_status": row["current_status"],
                    "primary_source_route": row["primary_source_route"],
                    "secondary_source_route": row["secondary_source_route"],
                    "collection_or_annotation_control": row["collection_or_annotation_control"],
                    "mandatory_lineage": row["mandatory_lineage"],
                    "admission_rule": row["admission_rule"],
                    "dependency_or_blocker": row["dependency_or_blocker"],
                    "local_recovery_status": "0_OF_1016_CONTENT_VALUES_RECOVERABLE",
                    "local_recovery_reason": "NO_ONE_TO_ONE_CUTOFF_SAFE_PRIMARY_OR_AUTHORIZED_PROVIDER_EVIDENCE_OBJECT",
                }
            )
    observed = [row["field_name"] for row in registry]
    assert set(observed) == set(EXPECTED_FEATURE_FIELDS), (observed, EXPECTED_FEATURE_FIELDS)
    assert len(observed) == 17 and len(set(observed)) == 17
    registry.sort(key=lambda row: EXPECTED_FEATURE_FIELDS.index(row["field_name"]))
    return registry


def make_schema() -> dict[str, Any]:
    slot_schema = {
        "type": "object",
        "additionalProperties": False,
        "required": [
            "source_sheet_state",
            "value",
            "collection_state",
            "local_recovery_state",
            "source_evidence_ref",
            "publication_at",
            "available_before_cutoff",
            "admission_state",
        ],
        "properties": {
            "source_sheet_state": {"const": "NEEDS_RESEARCH"},
            "value": {"type": "null"},
            "collection_state": {"const": "NOT_COLLECTED"},
            "local_recovery_state": {"const": "BLOCKED_EXTERNAL_HISTORICAL_RETRIEVAL_REQUIRED"},
            "source_evidence_ref": {"type": "null"},
            "publication_at": {"type": "null"},
            "available_before_cutoff": {"type": "null"},
            "admission_state": {"const": "BLOCKED_NO_LOCAL_CUTOFF_SAFE_EVIDENCE"},
        },
    }
    field_properties = {field: slot_schema for field in EXPECTED_FEATURE_FIELDS}
    return {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "$id": SCHEMA_ID,
        "title": "G3-E fail-closed annotation ingest queue row",
        "description": "Non-admitted authoring queue; null values are intentional and must not be coerced to zero/false.",
        "type": "object",
        "additionalProperties": False,
        "required": [
            "queue_id",
            "row_key",
            "source_workbook_row",
            "window_id",
            "company_no",
            "company_id",
            "canonical_name",
            "krx_code",
            "snapshot_cutoff_at",
            "entry_date",
            "historical_eligibility_status",
            "source_bundle_plan_state",
            "annotation_sidecar_state",
            "outcome_access_flag",
            "row_ingest_state",
            "field_registry_id",
            "field_slots",
        ],
        "properties": {
            "queue_id": {"const": QUEUE_ID},
            "row_key": {"type": "string", "minLength": 1},
            "source_workbook_row": {"type": "integer", "minimum": 2, "maximum": 1017},
            "window_id": {"enum": [f"W{i}" for i in range(1, 9)]},
            "company_no": {"type": "integer", "minimum": 1, "maximum": 127},
            "company_id": {"type": "string", "minLength": 1},
            "canonical_name": {"type": "string", "minLength": 1},
            "krx_code": {"type": "string", "minLength": 1},
            "snapshot_cutoff_at": {"type": "string", "pattern": "^[0-9]{4}-[0-9]{2}-[0-9]{2}$"},
            "entry_date": {"type": "string", "pattern": "^[0-9]{4}-[0-9]{2}-[0-9]{2}$"},
            "historical_eligibility_status": {"type": ["string", "null"]},
            "source_bundle_plan_state": {"const": "UNFROZEN"},
            "annotation_sidecar_state": {"const": "ABSENT"},
            "outcome_access_flag": {"type": "null"},
            "row_ingest_state": {"const": "QUEUE_ONLY_NOT_ADMITTED"},
            "field_registry_id": {"const": FIELD_REGISTRY_ID},
            "field_slots": {
                "type": "object",
                "additionalProperties": False,
                "required": EXPECTED_FEATURE_FIELDS,
                "properties": field_properties,
            },
        },
    }


def field_slot() -> dict[str, Any]:
    return {
        "source_sheet_state": "NEEDS_RESEARCH",
        "value": None,
        "collection_state": "NOT_COLLECTED",
        "local_recovery_state": "BLOCKED_EXTERNAL_HISTORICAL_RETRIEVAL_REQUIRED",
        "source_evidence_ref": None,
        "publication_at": None,
        "available_before_cutoff": None,
        "admission_state": "BLOCKED_NO_LOCAL_CUTOFF_SAFE_EVIDENCE",
    }


def output_hash_record(path: Path) -> dict[str, Any]:
    return {"path": path.name, "bytes": path.stat().st_size, "sha256": sha256(path)}


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)

    expected_hashes = {
        WORKBOOK: "44501584c9dc6224637e9193219c1e8c87507af77dc15dc3944a3d04af524cda",
        BUILD_MANIFEST: "5b78b6f0ea8cbdc2684e37724e3f0323ae8c50f1ae6dc1547e444d3e9c0eb7a1",
        WINDOWS: "96d63cc98a01b6332cf9486440e7f3fdaa0ec5a2d605f21bc14a4025b46e69fe",
        DENOMINATOR_QUEUE: "02bde437c04b1cc3d314b30e9bdd41bdb9a9164d0d2df4468728bdab8089eb62",
        MEMBERSHIP: "6a7c40b2a8bd52353a944f108dd556bf1dc05a520926aebb6d1bca4ae3b48f7c",
    }
    for path, expected in expected_hashes.items():
        assert sha256(path) == expected, (path, sha256(path), expected)

    windows = {row["window_id"]: row for row in read_csv(WINDOWS)}
    assert set(windows) == {f"W{i}" for i in range(1, 9)}

    wb = load_workbook(WORKBOOK, read_only=False, data_only=True)
    headers, rows = worksheet_rows(wb, "Thin_PIT")
    assert len(rows) == 1016
    assert all(field in headers for field in EXPECTED_FEATURE_FIELDS)

    row_keys: set[str] = set()
    window_counts: Counter[str] = Counter()
    feature_state_counts: dict[str, Counter[Any]] = {
        field: Counter(row[field] for row in rows) for field in EXPECTED_FEATURE_FIELDS
    }
    for field, counts in feature_state_counts.items():
        assert counts == Counter({"NEEDS_RESEARCH": 1016}), (field, counts)

    for row in rows:
        key = f'{row["window"]}|{row["company_id"]}'
        assert key not in row_keys
        row_keys.add(key)
        window_counts[str(row["window"])] += 1
        assert str(row["snapshot_cutoff_at"]) == windows[str(row["window"])]["snapshot_cutoff"]

    assert len(row_keys) == 1016
    assert window_counts == Counter({f"W{i}": 127 for i in range(1, 9)})
    assert Counter(row["evidence_status"] for row in rows) == Counter({"NOT_RESEARCHED": 1016})
    assert Counter(row["source_evidence_ref"] for row in rows) == Counter({None: 1016})
    assert Counter(row["publication_at"] for row in rows) == Counter({None: 1016})
    assert Counter(row["last_verified_observed_at"] for row in rows) == Counter({None: 1016})
    assert Counter(row["freshness_staleness"] for row in rows) == Counter({"UNKNOWN": 1016})
    assert Counter(row["thin_pit_slot_status"] for row in rows) == Counter({"INITIALIZED": 1016})
    assert Counter(row["thin_pit_completion_status"] for row in rows) == Counter({"INCOMPLETE": 1016})

    registry = manifest_registry()
    registry_path = OUT / "G3_E_ANNOTATION_INGEST_FIELD_REGISTRY_v0.1.json"
    registry_doc = {
        "field_registry_id": FIELD_REGISTRY_ID,
        "observed_at_kst": OBSERVED_AT_KST,
        "scope": "17 Thin_PIT dynamic feature/reference fields",
        "candidate_class": "NON_ADMITTED_AUTHORING_QUEUE",
        "field_count": len(registry),
        "fields": registry,
        "local_source_mapping": [
            {
                "source": WORKBOOK.name,
                "mapping": "Thin_PIT row identity/cutoff and existing field states",
                "recoverable": "1016_ROW_KEYS_AND_17272_QUEUE_SLOTS",
                "not_recoverable": "ANNOTATION_CONTENT_OR_EVIDENCE_LINEAGE",
            },
            {
                "source": "SEMI-EVAL-CORE / SEMI-DATA-ROUTE / SEMI-PIT-LEDGER",
                "mapping": "field taxonomy, routing, PIT cutoff, missingness semantics",
                "recoverable": "SCHEMA_AND_CONTROL_RULES",
                "not_recoverable": "ROW_LEVEL_HISTORICAL_VALUES",
            },
            {
                "source": "SEMI-COMPANY-MASTER / U81_F1",
                "mapping": "static F1 structural context",
                "recoverable": "F1_CONTEXT_ONLY",
                "not_recoverable": "DYNAMIC_QUALIFICATION_DESIGN_WIN_REPEAT_ORDER_OR_FAB_CAPEX_STATE",
            },
            {
                "source": "Historical_BP / Listing_Tradability / Identity_Ledger / Corporate_Actions",
                "mapping": "eligibility, identity, or CA domains",
                "recoverable": "CROSS_DOMAIN_METADATA_ONLY",
                "not_recoverable": "G3_E_17_FIELD_ANNOTATION_VALUES",
            },
            {
                "source": "marcap-2024/2025/2026 Parquet components",
                "mapping": "price and market-cap domain",
                "recoverable": "PRICE_FIELDS_ALREADY_SEPARATELY_POPULATED",
                "not_recoverable": "VALUATION_WITHOUT_PIT_DENOMINATOR_OR_ANY_NONPRICE_ANNOTATION",
            },
            {
                "source": "2026-08-09 through 2026-08-14 research documents",
                "mapping": "derived research context and source-route hints",
                "recoverable": "NO_ADMITTED_COMPANY_WINDOW_ANNOTATION",
                "not_recoverable": "HISTORICAL_BACKFILL_BECAUSE_DOCUMENTS_POSTDATE_ALL_W1_W8_SNAPSHOT_CUTOFFS_AND_LACK_ONE_TO_ONE_PRIMARY_SOURCE_OBJECTS",
            },
        ],
        "prohibited_shortcuts": [
            "QUALIFICATION_BARRIER_TO_QUALIFICATION_STATUS",
            "FAB_EXPOSURE_TO_FAB_CAPEX_STATE_STATUS",
            "CUSTOMER_STRUCTURE_TO_DESIGN_WIN_OR_CUSTOMER_ADOPTION",
            "PRICE_OR_MARKET_CAP_TO_VALUATION_WITHOUT_PIT_EARNINGS_DENOMINATOR",
            "POST_CUTOFF_PROJECT_REPORT_TO_HISTORICAL_PUBLICATION_AT",
            "MISSING_TO_ZERO_FALSE_OR_NEGATIVE_BUSINESS_FACT",
        ],
    }
    registry_path.write_text(json.dumps(registry_doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    schema_path = OUT / "G3_E_ANNOTATION_INGEST_SCHEMA_v0.1.json"
    schema_path.write_text(json.dumps(make_schema(), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    queue_path = OUT / "G3_E_ANNOTATION_INGEST_QUEUE_v0.1.jsonl"
    with queue_path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            window_id = str(row["window"])
            krx_code = str(row["KRX_code"])
            queue_row = {
                "queue_id": QUEUE_ID,
                "row_key": f'{window_id}|{row["company_id"]}',
                "source_workbook_row": int(row["_workbook_row"]),
                "window_id": window_id,
                "company_no": int(row["company_no"]),
                "company_id": str(row["company_id"]),
                "canonical_name": str(row["canonical_name"]),
                "krx_code": krx_code,
                "snapshot_cutoff_at": str(row["snapshot_cutoff_at"]),
                "entry_date": windows[window_id]["entry_date"],
                "historical_eligibility_status": row["historical_eligibility_status"],
                "source_bundle_plan_state": "UNFROZEN",
                "annotation_sidecar_state": "ABSENT",
                "outcome_access_flag": None,
                "row_ingest_state": "QUEUE_ONLY_NOT_ADMITTED",
                "field_registry_id": FIELD_REGISTRY_ID,
                "field_slots": {field: field_slot() for field in EXPECTED_FEATURE_FIELDS},
            }
            handle.write(json.dumps(queue_row, ensure_ascii=False, separators=(",", ":")) + "\n")

    def non_null_count(sheet: str, column: str) -> int:
        _, sheet_rows = worksheet_rows(wb, sheet)
        return sum(row[column] not in (None, "") for row in sheet_rows)

    cross_domain_dates = {
        "Identity_Ledger.publication_at": non_null_count("Identity_Ledger", "publication_at"),
        "U81_F1.publication_at": non_null_count("U81_F1", "publication_at"),
        "Historical_BP.publication_at": non_null_count("Historical_BP", "publication_at"),
        "Corporate_Actions.evidence_publication_at": non_null_count("Corporate_Actions", "evidence_publication_at"),
    }
    wb.close()

    source_paths = [WORKBOOK, BUILD_MANIFEST, WINDOWS, DENOMINATOR_QUEUE, MEMBERSHIP]
    source_paths.extend(ROOT / "project_sources" / name for name in POST_CUTOFF_CONTEXT_DOCS)
    source_paths.extend(
        [
            ROOT / "project_sources/10-SEMI-PRICE-LEDGER_IMPORT-MANIFEST_v1.0.csv",
            ROOT / "project_sources/13-SEMI-PRICE-LEDGER_v1.0_SCHEMA.csv",
            ROOT / "project_sources/15-SEMI-PRICE-LEDGER_v1.0_INGEST-AUDIT_BLOCKED_2026-08-14.csv",
            ROOT / "project_sources/16-marcap-2025.parquet",
        ]
    )
    input_manifest = [
        {"path": str(path), "bytes": path.stat().st_size, "sha256": sha256(path)}
        for path in source_paths
    ]

    summary_path = OUT / "G3_E_LOCAL_RECOVERABILITY_SUMMARY_v0.1.json"
    summary = {
        "artifact_id": "AAA-M3TOP3-G3E-LOCAL-RECOVERABILITY-20260826-0125",
        "observed_at_kst": OBSERVED_AT_KST,
        "execution_role": "NON_VALIDATOR_WORKER",
        "validator_hold": True,
        "global_validation": False,
        "validation_loop": False,
        "source_mutation": False,
        "git_issue_mutation": False,
        "gate_effect": "NONE",
        "decision": "NO_ANNOTATION_CONTENT_VALUE_IS_DETERMINISTICALLY_RECOVERABLE_FROM_LOCAL_INPUTS",
        "counts": {
            "thin_pit_rows": 1016,
            "unique_company_window_keys": 1016,
            "companies": 127,
            "windows": 8,
            "annotation_fields": 17,
            "annotation_field_slots": 17272,
            "queue_rows_recoverable": 1016,
            "queue_slots_recoverable": 17272,
            "content_values_recoverable": 0,
            "publication_at_recoverable_for_annotation": 0,
            "source_evidence_ref_recoverable_for_annotation": 0,
            "available_before_cutoff_determinable": 0,
            "admitted_annotation_rows": 0,
            "annotation_sidecar_rows": 0,
        },
        "window_counts": dict(sorted(window_counts.items())),
        "feature_state_counts": {
            field: {str(key): value for key, value in counts.items()}
            for field, counts in feature_state_counts.items()
        },
        "cross_domain_publication_metadata": {
            "counts": cross_domain_dates,
            "annotation_admissibility": "ZERO; dates are bound to identity/F1/eligibility/CA records, not to the 17 dynamic feature slots",
        },
        "mechanically_recoverable": [
            "1016 exact company-window row keys",
            "127-by-8 coverage and source workbook row numbers",
            "snapshot cutoffs and candidate entry-date joins",
            "17-field source-route/control registry",
            "17272 fail-closed NOT_COLLECTED queue slots",
            "exact ingest JSON Schema",
        ],
        "blocked": [
            "historical evidence retrieval",
            "one-to-one source object identity and source hash",
            "feature-level publication_at and observation period",
            "available-before-cutoff determination",
            "frozen outcome-concealed source bundle",
            "annotation/access sidecar and independent interpretation lineage",
            "exact-v1 feature admission semantics while G1 identity remains unresolved",
        ],
        "input_manifest": input_manifest,
        "output_manifest": [
            output_hash_record(registry_path),
            output_hash_record(schema_path),
            output_hash_record(queue_path),
        ],
        "targeted_mechanical_checks": {
            "input_hash_bindings": "PASS_5_EXACT",
            "thin_pit_row_count": "PASS_1016",
            "row_key_uniqueness": "PASS_1016_OF_1016",
            "window_balance": "PASS_127_EACH_W1_W8",
            "feature_field_count": "PASS_17",
            "feature_source_state": "PASS_ALL_17272_NEEDS_RESEARCH",
            "evidence_status": "PASS_ALL_1016_NOT_RESEARCHED",
            "publication_and_source_refs": "PASS_ALL_1016_NULL",
            "queue_state": "PASS_FAIL_CLOSED_NO_VALUE_INVENTED",
        },
        "ewu_recommendation": {
            "unit": "G3-E annotation dependency",
            "earned_ewu": 1,
            "basis": "Executable fail-closed schema/registry/1016-row queue produced; no content evidence or gate closure",
            "ceiling": "Do not award the remaining G3-E content-closure EWU",
        },
        "remaining_worker_only_timing": {
            "package_and_readback_p50_hours": [0.25, 0.5],
            "package_and_readback_p90_hours": [0.75, 1.0],
            "true_annotation_retrieval_and_dual_coding_eta": "NOT_MEASURABLE_FROM_CURRENT_LOCAL_INPUTS",
            "workload_floor_at_one_minute_per_company_window_hours": 16.9,
        },
        "claim_ceiling": [
            "NO_NEW_PASS",
            "NO_G3_GATE_CLOSURE",
            "NO_ANNOTATION_TRUTH_CLAIM",
            "NO_PUBLICATION_DATE_INVENTION",
            "NO_CURRENT_OR_POST_CUTOFF_BACKFILL",
            "NO_SCORE_GOLDEN_REPLAY_EOPT_OR_PRODUCTION_CLAIM",
        ],
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
